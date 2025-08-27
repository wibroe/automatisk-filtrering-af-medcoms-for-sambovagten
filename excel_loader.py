"""
Module for loading data from Excel files.
"""
import logging
from typing import List

import openpyxl


def load_indsatser_list(excel_file_path: str) -> List[str]:
    """
    Load indsatser (services) list from Excel file.
    
    Args:
        excel_file_path: Path to the Excel file
        
    Returns:
        List of service names from the first column
    """
    logger = logging.getLogger(__name__)
    logger.info(f"Loading indsatser list from: {excel_file_path}")
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file_path)
        
        # Get the 'Liste' worksheet
        if 'Liste' not in workbook.sheetnames:
            raise ValueError(f"Worksheet 'Liste' not found in {excel_file_path}. Available sheets: {workbook.sheetnames}")
        
        worksheet = workbook['Liste']
        
        # Create list for service names
        indsatser_list = []
        
        # Skip header row (row 1) and iterate through data rows
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # First column has a value
                service_name = str(row[0]).strip()
                if service_name:  # Not empty after stripping
                    indsatser_list.append(service_name)
        
        logger.info(f"Loaded {len(indsatser_list)} indsatser")
        
        return indsatser_list
        
    except Exception as e:
        logger.error(f"Error loading indsatser list: {e}")
        raise
